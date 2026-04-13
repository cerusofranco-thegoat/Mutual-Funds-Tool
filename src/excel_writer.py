"""Generate formatted Excel output with per-fund sheets and an Analysis sheet."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from .config import Config
from .models import AnalyticalReport, FundData

# --- Style constants ---

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

GO_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NOGO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
CONDITIONAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

SUITABLE_FILL = GO_FILL
MARGINAL_FILL = CONDITIONAL_FILL
UNSUITABLE_FILL = NOGO_FILL


def _sanitize_sheet_name(name: str) -> str:
    """Create a valid Excel sheet name (max 31 chars, no special chars)."""
    clean = re.sub(r'[\\/*?\[\]:]', '_', name)
    return clean[:31]


def _apply_header_row(ws, row: int, values: list[str], start_col: int = 1) -> None:
    """Apply header styling to a row."""
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_data_cell(ws, row: int, col: int, value, number_format: str | None = None) -> None:
    """Write a data cell with border and optional number format."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(wrap_text=True)
    if number_format:
        cell.number_format = number_format


def _decision_fill(decision: str) -> PatternFill | None:
    """Return fill color based on decision/status text."""
    d = decision.lower().strip()
    if d in ("go", "approved", "suitable", "pass"):
        return GO_FILL
    elif d in ("no-go", "nogo", "rejected", "unsuitable", "fail"):
        return NOGO_FILL
    elif d in ("conditional", "marginal", "warning"):
        return CONDITIONAL_FILL
    return None


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width


# --- Fund sheet writer ---

def _write_fund_sheet(ws, fund: FundData) -> None:
    """Write all data for a single fund into a worksheet."""
    row = 1
    fid = fund.identification

    # Title
    ws.cell(row=row, column=1, value=fid.fund_name).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    # Fund info
    info = [
        ("Date", fid.date),
        ("Fund Type", fid.fund_type),
        ("Company", fid.company),
        ("Source File", fund.source_file),
    ]
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=row, column=2, value=value or "N/A")
        row += 1

    row += 1

    # --- Value at Risk ---
    ws.cell(row=row, column=1, value="Value at Risk (VaR)").font = SECTION_FONT
    row += 1

    if fund.risks and fund.risks.var:
        var = fund.risks.var
        _apply_header_row(ws, row, ["VaR Value (%)", "Period", "Confidence Level (%)"])
        row += 1
        _apply_data_cell(ws, row, 1, var.value, "0.00")
        _apply_data_cell(ws, row, 2, var.period)
        _apply_data_cell(ws, row, 3, var.confidence_level, "0.00")
        row += 1
    else:
        ws.cell(row=row, column=1, value="No VaR data available")
        row += 1

    row += 1

    # --- Risk Summary ---
    ws.cell(row=row, column=1, value="Risk Summary").font = SECTION_FONT
    row += 1

    if fund.risks and fund.risks.risk_items:
        _apply_header_row(ws, row, ["Category", "Description", "Severity"])
        row += 1
        for item in fund.risks.risk_items:
            _apply_data_cell(ws, row, 1, item.category)
            _apply_data_cell(ws, row, 2, item.description)
            _apply_data_cell(ws, row, 3, item.severity or "N/A")
            fill = _decision_fill(item.severity or "")
            if fill:
                ws.cell(row=row, column=3).fill = fill
            row += 1

        if fund.risks.risk_summary:
            row += 1
            ws.cell(row=row, column=1, value="Overall Summary").font = SUBHEADER_FONT
            row += 1
            ws.cell(row=row, column=1, value=fund.risks.risk_summary)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No risk data available")
        row += 1

    row += 1

    # --- Financial Returns ---
    ws.cell(row=row, column=1, value="Financial Returns (%)").font = SECTION_FONT
    row += 1

    if fund.returns and fund.returns.series:
        # Collect all unique periods across all series
        all_periods: list[str] = []
        for s in fund.returns.series:
            for r in s.returns:
                if r.period not in all_periods:
                    all_periods.append(r.period)

        headers = ["Series / Class"] + all_periods
        _apply_header_row(ws, row, headers)
        row += 1

        for series in fund.returns.series:
            _apply_data_cell(ws, row, 1, series.series_name)
            returns_map = {r.period: r.value for r in series.returns}
            for i, period in enumerate(all_periods):
                val = returns_map.get(period)
                _apply_data_cell(ws, row, 2 + i, val, "0.00" if val is not None else None)
            row += 1

        # Benchmarks
        if fund.returns.benchmarks:
            for bm in fund.returns.benchmarks:
                _apply_data_cell(ws, row, 1, f"Benchmark: {bm.benchmark_name}")
                ws.cell(row=row, column=1).font = Font(italic=True)
                returns_map = {r.period: r.value for r in bm.returns}
                for i, period in enumerate(all_periods):
                    val = returns_map.get(period)
                    _apply_data_cell(ws, row, 2 + i, val, "0.00" if val is not None else None)
                row += 1
    else:
        ws.cell(row=row, column=1, value="No returns data available")
        row += 1

    row += 1

    # --- Investment Portfolio ---
    ws.cell(row=row, column=1, value="Investment Portfolio").font = SECTION_FONT
    row += 1

    if fund.portfolio and fund.portfolio.holdings:
        portfolio_headers = [
            "Asset Name", "Type", "Issuer", "% Portfolio",
            "Market Value", "Currency", "Maturity", "Rating", "Coupon %",
        ]
        _apply_header_row(ws, row, portfolio_headers)
        row += 1

        for h in fund.portfolio.holdings:
            _apply_data_cell(ws, row, 1, h.asset_name)
            _apply_data_cell(ws, row, 2, h.asset_type or "")
            _apply_data_cell(ws, row, 3, h.issuer or "")
            _apply_data_cell(ws, row, 4, h.percentage, "0.00")
            _apply_data_cell(ws, row, 5, h.market_value, "#,##0.00")
            _apply_data_cell(ws, row, 6, h.currency or "")
            _apply_data_cell(ws, row, 7, h.maturity_date or "")
            _apply_data_cell(ws, row, 8, h.credit_rating or "")
            _apply_data_cell(ws, row, 9, h.coupon_rate, "0.00" if h.coupon_rate else None)
            row += 1

        if fund.portfolio.total_assets:
            row += 1
            ws.cell(row=row, column=1, value="Total Net Assets").font = SUBHEADER_FONT
            _apply_data_cell(ws, row, 5, fund.portfolio.total_assets, "#,##0.00")
            _apply_data_cell(ws, row, 6, fund.portfolio.total_assets_currency or "")
    else:
        ws.cell(row=row, column=1, value="No portfolio data available")

    _auto_width(ws)
    ws.freeze_panes = "A2"


# --- Analysis sheet writer ---

def _write_analysis_sheet(ws, report: AnalyticalReport) -> None:
    """Write the cross-fund analytical report into a worksheet."""
    row = 1

    ws.cell(row=row, column=1, value="Mutual Fund Analytical Report").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Section 1: Asset Distribution
    ws.cell(row=row, column=1, value="1. Asset Distribution").font = SECTION_FONT
    row += 1
    if report.asset_distribution:
        _apply_header_row(ws, row, ["Asset Type", "Count", "Total Value", "% of All"])
        row += 1
        for ad in report.asset_distribution:
            _apply_data_cell(ws, row, 1, ad.asset_type)
            _apply_data_cell(ws, row, 2, ad.count)
            _apply_data_cell(ws, row, 3, ad.total_value, "#,##0.00")
            _apply_data_cell(ws, row, 4, ad.percentage, "0.00")
            row += 1
    row += 1

    # Section 2: Fund Investment Summary
    ws.cell(row=row, column=1, value="2. Fund Investment Summary").font = SECTION_FONT
    row += 1
    if report.fund_summaries:
        _apply_header_row(ws, row, ["Fund Name", "Total AUM", "Currency", "# Holdings", "Avg Return 1Y (%)"])
        row += 1
        for fs in report.fund_summaries:
            _apply_data_cell(ws, row, 1, fs.fund_name)
            _apply_data_cell(ws, row, 2, fs.total_aum, "#,##0.00")
            _apply_data_cell(ws, row, 3, fs.aum_currency or "")
            _apply_data_cell(ws, row, 4, fs.num_holdings)
            _apply_data_cell(ws, row, 5, fs.avg_return_1y, "0.00")
            row += 1
    row += 1

    # Section 3: Go/No-Go Decisions
    ws.cell(row=row, column=1, value="3. Go / No-Go Decisions").font = SECTION_FONT
    row += 1
    if report.go_nogo_decisions:
        _apply_header_row(ws, row, ["Fund Name", "Decision", "Justification"])
        row += 1
        for d in report.go_nogo_decisions:
            _apply_data_cell(ws, row, 1, d.fund_name)
            _apply_data_cell(ws, row, 2, d.decision)
            fill = _decision_fill(d.decision)
            if fill:
                ws.cell(row=row, column=2).fill = fill
            _apply_data_cell(ws, row, 3, d.justification)
            row += 1
    row += 1

    # Section 4: Risk Tolerance Alignment
    ws.cell(row=row, column=1, value="4. Risk Tolerance Alignment").font = SECTION_FONT
    row += 1
    if report.risk_tolerance:
        _apply_header_row(ws, row, ["Fund Name", "Conservative", "Moderate", "Aggressive", "Rationale"])
        row += 1
        for rt in report.risk_tolerance:
            _apply_data_cell(ws, row, 1, rt.fund_name)
            for col_idx, val in [(2, rt.conservative), (3, rt.moderate), (4, rt.aggressive)]:
                _apply_data_cell(ws, row, col_idx, val)
                fill = _decision_fill(val)
                if fill:
                    ws.cell(row=row, column=col_idx).fill = fill
            _apply_data_cell(ws, row, 5, rt.rationale or "")
            row += 1
    row += 1

    # Section 5: Capital Sizing
    ws.cell(row=row, column=1, value="5. Capital Sizing Recommendations").font = SECTION_FONT
    row += 1
    if report.capital_sizing:
        _apply_header_row(ws, row, ["Fund Name", "Recommended %", "Min %", "Max %", "Rationale"])
        row += 1
        for cs in report.capital_sizing:
            _apply_data_cell(ws, row, 1, cs.fund_name)
            _apply_data_cell(ws, row, 2, cs.recommended_pct, "0.00")
            _apply_data_cell(ws, row, 3, cs.min_pct, "0.00")
            _apply_data_cell(ws, row, 4, cs.max_pct, "0.00")
            _apply_data_cell(ws, row, 5, cs.rationale or "")
            row += 1
    row += 1

    # Section 6: Diversification Checks
    ws.cell(row=row, column=1, value="6. Portfolio Diversification Checks").font = SECTION_FONT
    row += 1
    if report.diversification_checks:
        _apply_header_row(ws, row, ["Metric", "Status", "Details"])
        row += 1
        for dc in report.diversification_checks:
            _apply_data_cell(ws, row, 1, dc.metric)
            _apply_data_cell(ws, row, 2, dc.status)
            fill = _decision_fill(dc.status)
            if fill:
                ws.cell(row=row, column=2).fill = fill
            _apply_data_cell(ws, row, 3, dc.details)
            row += 1
    row += 1

    # Section 7: Investment Horizon
    ws.cell(row=row, column=1, value="7. Investment Horizon Fit").font = SECTION_FONT
    row += 1
    if report.investment_horizons:
        _apply_header_row(ws, row, ["Fund Name", "Horizon", "Rationale"])
        row += 1
        for ih in report.investment_horizons:
            _apply_data_cell(ws, row, 1, ih.fund_name)
            _apply_data_cell(ws, row, 2, ih.horizon)
            _apply_data_cell(ws, row, 3, ih.rationale)
            row += 1
    row += 1

    # Section 8: Value Assessment
    ws.cell(row=row, column=1, value="8. Value Assessment").font = SECTION_FONT
    row += 1
    if report.value_assessments:
        _apply_header_row(ws, row, ["Fund Name", "Fee Level", "Return/Risk", "Fair Value?", "Rationale"])
        row += 1
        for va in report.value_assessments:
            _apply_data_cell(ws, row, 1, va.fund_name)
            _apply_data_cell(ws, row, 2, va.fee_level or "")
            _apply_data_cell(ws, row, 3, va.return_risk_ratio or "")
            _apply_data_cell(ws, row, 4, va.fair_value)
            fill = _decision_fill(va.fair_value)
            if fill:
                ws.cell(row=row, column=4).fill = fill
            _apply_data_cell(ws, row, 5, va.rationale or "")
            row += 1
    row += 1

    # Section 9: Fund Usage Approval
    ws.cell(row=row, column=1, value="9. Fund Usage Approval").font = SECTION_FONT
    row += 1
    if report.fund_approvals:
        _apply_header_row(ws, row, ["Fund Name", "Status", "Conditions"])
        row += 1
        for fa in report.fund_approvals:
            _apply_data_cell(ws, row, 1, fa.fund_name)
            _apply_data_cell(ws, row, 2, fa.status)
            fill = _decision_fill(fa.status)
            if fill:
                ws.cell(row=row, column=2).fill = fill
            _apply_data_cell(ws, row, 3, fa.conditions or "")
            row += 1
    row += 1

    # Overall summary
    if report.overall_summary:
        ws.cell(row=row, column=1, value="Overall Summary").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value=report.overall_summary)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    _auto_width(ws)
    ws.freeze_panes = "A2"


# --- Main workbook generator ---

def generate_workbook(
    config: Config,
    funds: list[FundData],
    report: AnalyticalReport | None,
) -> Path:
    """Generate the complete Excel output file."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create per-fund sheets
    for fund in funds:
        sheet_name = _sanitize_sheet_name(fund.identification.fund_name)
        # Ensure unique sheet name
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{len(existing)}"
        ws = wb.create_sheet(title=sheet_name)
        _write_fund_sheet(ws, fund)

    # Create analysis sheet
    if report:
        ws_analysis = wb.create_sheet(title="Analysis")
        _write_analysis_sheet(ws_analysis, report)
        # Move analysis to first position
        wb.move_sheet("Analysis", offset=-len(wb.sheetnames) + 1)

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = config.output_dir / f"blackrock_analysis_{timestamp}.xlsx"
    config.output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path
