"""Read supplementary Excel input files into normalized DocumentContent."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from .models import DocumentContent, PageContent

logger = logging.getLogger(__name__)


def _page_from_rows(index: int, sheet_name: str, rows: list[list[str]]) -> PageContent:
    text_lines = ["\t".join(c for c in r if c) for r in rows if any(r)]
    text = f"[Sheet: {sheet_name}]\n" + "\n".join(text_lines)
    return PageContent(
        page_number=index + 1,
        text=text,
        tables=[rows] if rows else [],
    )


def _extract_xls_legacy(file_path: Path) -> list[PageContent]:
    import xlrd  # type: ignore

    pages: list[PageContent] = []
    book = xlrd.open_workbook(str(file_path))
    for i, sheet in enumerate(book.sheets()):
        rows: list[list[str]] = []
        for r in range(sheet.nrows):
            row = sheet.row_values(r)
            rows.append([str(c) if c != "" else "" for c in row])
        pages.append(_page_from_rows(i, sheet.name, rows))
    return pages


def _looks_like_spreadsheetml(file_path: Path) -> bool:
    """Detect Excel 2003 XML (SpreadsheetML) files mislabeled as .xls."""
    try:
        with open(file_path, "rb") as f:
            head = f.read(512)
    except OSError:
        return False
    head = head.lstrip(b"\xef\xbb\xbf").lstrip()
    return head.startswith(b"<?xml") and b"urn:schemas-microsoft-com:office:spreadsheet" in head[:512] or (
        head.startswith(b"<?xml") and b"<Workbook" in head
    )


def _extract_spreadsheetml(file_path: Path) -> list[PageContent]:
    """Parse Excel 2003 XML (SpreadsheetML) workbooks."""
    import xml.etree.ElementTree as ET

    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    tree = ET.parse(file_path)
    root = tree.getroot()

    pages: list[PageContent] = []
    for i, worksheet in enumerate(root.findall("ss:Worksheet", ns)):
        sheet_name = worksheet.get(f"{{{ns['ss']}}}Name", f"Sheet{i+1}")
        rows: list[list[str]] = []
        table = worksheet.find("ss:Table", ns)
        if table is None:
            pages.append(_page_from_rows(i, sheet_name, rows))
            continue

        for row in table.findall("ss:Row", ns):
            cells_out: list[str] = []
            col_idx = 0
            for cell in row.findall("ss:Cell", ns):
                index_attr = cell.get(f"{{{ns['ss']}}}Index")
                if index_attr:
                    target = int(index_attr) - 1
                    while col_idx < target:
                        cells_out.append("")
                        col_idx += 1
                data = cell.find("ss:Data", ns)
                cells_out.append(data.text if data is not None and data.text else "")
                col_idx += 1
            rows.append(cells_out)

        pages.append(_page_from_rows(i, sheet_name, rows))
    return pages


def _extract_xlsx(file_path: Path) -> list[PageContent]:
    pages: list[PageContent] = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        pages.append(_page_from_rows(i, sheet_name, rows))
    wb.close()
    return pages


def extract_excel(file_path: Path) -> DocumentContent:
    """Read an Excel file and convert each sheet into a PageContent."""
    pages: list[PageContent] = []
    try:
        if _looks_like_spreadsheetml(file_path):
            pages = _extract_spreadsheetml(file_path)
        elif file_path.suffix.lower() == ".xls":
            pages = _extract_xls_legacy(file_path)
        else:
            pages = _extract_xlsx(file_path)
    except Exception as e:
        logger.error("Failed to read Excel file %s: %s", file_path.name, e)

    return DocumentContent(filename=file_path.name, pages=pages)
